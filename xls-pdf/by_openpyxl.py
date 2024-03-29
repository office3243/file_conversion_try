# XLSXtoPDF.py

# Program to convert the data from an XLSX file to PDF.
# Uses the openpyxl library and xtopdf.

# Author: Vasudev Ram - http://jugad2.blogspot.com
# Copyright 2015 Vasudev Ram.

from openpyxl import load_workbook
from pdfrw import PdfWriter


inp = "C:/Users/eWay/Desktop/aamer/Sample-Sales-Data.xlsx"
out = "C:/Users/eWay/Desktop/aamer/Sample-Sales-Data.pdf"

workbook = load_workbook(inp, guess_types=True, data_only=True)
worksheet = workbook.active

pw = PdfWriter(out)
pw.setFont('Courier', 12)
pw.setHeader('XLSXtoPDF.py - convert XLSX data to PDF')
pw.setFooter('Generated using openpyxl and xtopdf')

ws_range = worksheet.iter_rows('A1:H13')
for row in ws_range:
    s = ''
    for cell in row:
        if cell.value is None:
            s += ' ' * 11
        else:
            s += str(cell.value).rjust(10) + ' '
    pw.writeLine(s)
pw.savePage()
pw.close()
